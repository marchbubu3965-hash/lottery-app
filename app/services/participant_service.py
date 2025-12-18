import sqlite3
from typing import List, Tuple

from app.db.database import get_connection


class ParticipantService:
    """
    名單管理服務：
    - CRUD
    - Excel 匯入
    - 抽籤狀態控制
    """

    # =========================
    # 基本 CRUD
    # =========================
    # def add_participant(self, name: str) -> None:
    #     if not name:
    #         raise ValueError("候選人名稱不可為空")

    #     conn = get_connection()
    #     cursor = conn.cursor()

    #     cursor.execute(
    #         "INSERT INTO participants (name) VALUES (?)",
    #         (name,)
    #     )

    #     conn.commit()
    #     conn.close()

    # def update_participant(self, participant_id: int, new_name: str) -> None:
    #     if not new_name:
    #         raise ValueError("新名稱不可為空")

    #     conn = get_connection()
    #     cursor = conn.cursor()

    #     cursor.execute(
    #         "UPDATE participants SET name = ? WHERE id = ?",
    #         (new_name, participant_id)
    #     )

    #     conn.commit()
    #     conn.close()

    # def delete_participant(self, participant_id: int) -> None:
    #     conn = get_connection()
    #     cursor = conn.cursor()

    #     cursor.execute(
    #         "DELETE FROM participants WHERE id = ?",
    #         (participant_id,)
    #     )

    #     conn.commit()
    #     conn.close()

    # =========================
    # 新增
    # =========================
    def add(self, name, employee_no=None):
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO participants (name, employee_no)
            VALUES (?, ?)
        """, (name, employee_no))

        conn.commit()
        conn.close()

    # =========================
    # 更新
    # =========================
    def update(self, pid, name, employee_no):
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            UPDATE participants
            SET name = ?, employee_no = ?
            WHERE id = ?
        """, (name, employee_no, pid))

        conn.commit()
        conn.close()

    # =========================
    # 刪除
    # =========================
    def delete(self, pid):
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            DELETE FROM participants
            WHERE id = ?
        """, (pid,))

        conn.commit()
        conn.close()

    def get_all_participants(self):
        conn = get_connection()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("""
            SELECT
                id,
                name,
                employee_no,
                is_active,
                created_at
            FROM participants
            ORDER BY id
        """)

        rows = cur.fetchall()
        conn.close()
        return rows

    def set_active(self, participant_id: int, active: bool):
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            UPDATE participants
            SET is_active = ?
            WHERE id = ?
        """, (1 if active else 0, participant_id))

        conn.commit()
        conn.close()

    # =========================
    # 抽籤相關
    # =========================
    def get_active_participants(self) -> List[Tuple]:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT id, name FROM participants WHERE is_active = 1"
        )
        rows = cursor.fetchall()

        conn.close()
        return rows

    def mark_as_selected(self, participant_id: int) -> None:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            "UPDATE participants SET is_active = 0 WHERE id = ?",
            (participant_id,)
        )

        conn.commit()
        conn.close()

    def reset_all_participants(self) -> None:
        """
        特別獎用：全部名單重設為可抽
        """
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            "UPDATE participants SET is_active = 1"
        )

        conn.commit()
        conn.close()

    # =========================
    # Excel 匯入
    # =========================
    def import_from_excel(self, file_path: str) -> int:
    
    # 從 Excel 匯入 participants
    # Excel 欄位：
    # | name | employee_no |
    
        try:
            import openpyxl
        except ImportError:
            raise ImportError("請先安裝 openpyxl")

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        conn = get_connection()
        cursor = conn.cursor()

        count = 0

        # 讀取 header
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = {h: i for i, h in enumerate(headers) if h}

        if "name" not in header_map:
            raise ValueError("Excel 必須包含 name 欄位")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[header_map["name"]]
            employee_no = row[header_map.get("employee_no")]

            if not name:
                continue

            cursor.execute(
                """
                INSERT INTO participants (name, employee_no)
                VALUES (?, ?)
                """,
                (str(name).strip(), str(employee_no).strip() if employee_no else None)
            )
            count += 1

        conn.commit()
        conn.close()
        return count

